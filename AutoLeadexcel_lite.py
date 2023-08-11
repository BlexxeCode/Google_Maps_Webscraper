import selenium.common.exceptions
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl import Workbook
import pandas as pd
from datetime import date
import time

#future additions
#  prompt to input maps url, business type, amount of businesses
#  turn into executable


browser = 'C://Users//andki//AppData//Local//Vivaldi//Application//vivaldi.exe'

path = 'C://Users//andki//OneDrive//Documents//EGTLeads.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb.active

options = webdriver.ChromeOptions()
options.binary_location = browser

Type = input('Type of locations?')
url2 = input("Google Maps URL")
Max = int(input('Amount of locations to obtain info from?'))
start = int(input("Row to start at?"))

driver = webdriver.Chrome('C://Users//andki//OneDrive//Documents//chromedriver_win32//chromedriver.exe', options=options)


url1 = 'https://chrome.google.com/webstore/detail/ublock-origin/cjpalhdlnbpafiamejdnhcphjbkeiagm'
driver.get(url1)
time.sleep(5)
driver.find_element('css selector', 'div.g-c-Hf').click()

time.sleep(10)
driver.get(url2)
time.sleep(8)

num = start


p = 1
while p <= Max:
    list = driver.find_elements('css selector', 'div.dbg0pd')
    print(len(list))
    for bis in list:
        if bis.text == '':
            continue
        bis.click()
        time.sleep(3)
        try:
            name = driver.find_element('css selector', 'div.SPZz6b')
            print(name.text)

            address = driver.find_element('css selector', 'span.LrzXr')


            spl_address = address.text.split(',')

            phone = driver.find_element('css selector', 'a.Od1FEc.mI8Pwc').get_attribute('data-phone-number')


            webtest = driver.find_elements('css selector', 'a.mI8Pwc')
            if len(webtest) == 2:
                web = webtest[-2].get_attribute('href')
            else:
                web = 'N/A'
        except selenium.common.exceptions.NoSuchElementException as e:
            continue
        try:
            Btype = driver.find_element('css selector', 'span.YhemCb').text
        except selenium.common.exceptions.NoSuchElementException as e:
            Btype = Type

        ws['F' + str(num)].value = str(date.today())
        ws['G' + str(num)].value = 'Lead'
        ws['H' + str(num)].value = ' '
        ws['I' + str(num)].value = phone
        ws['J' + str(num)].value = ''
        ws['K' + str(num)].value = Btype
        ws['L' + str(num)].value = 'Google Maps'
        ws['M' + str(num)].value = name.text
        ws['N' + str(num)].value = web
        ws['O' + str(num)].value = spl_address[0]
        if len(spl_address) < 2:
            ws['P' + str(num)].value = spl_address[0]
        else:
            ws['P' + str(num)].value = spl_address[-2]
        ws['Q' + str(num)].value = 'FL'
        ws['R' + str(num)].value = int(address.text[-5:])

        wb.save(path)

        print(Btype)
        print(address.text)
        print(web)
        print(phone)
        print('-' * 30)
        num += 1
        p += 1
        if p > Max:
            break
    driver.find_element('css selector', 'a#pnnext').click()
    time.sleep(5)

